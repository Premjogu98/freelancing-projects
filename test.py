# main = """Produced by
# elementz & jae5
# written by
# elementz, jae5 & burna boy
# additional vocals
# jae5, keven wolfshon, kwande bawa, paul bohumil goller & uncle t
# assistant mixing engineer
# mixgiant & joe begalla
# mixing engineer
# jesse ray ernster
# mastering engineer
# gerhard west phalen
# recording engineer
# eric issac utere
# a&r administrator
# irene sourlis
# a&r coordinator
# spaceship collective & matthew baus
# copyright ©
# on a spaceship records, bad habit records & atlantic records
# phonographic copyright ℗
# on a spaceship records, bad habit records & atlantic records
# label
# on a spaceship records, bad habit records & atlantic records
# distributor
# atlantic records
# release date
# july 8, 2022"""
# info_list = main.split("\n")
# print(info_list)
# while len(info_list) != 0:
#     print(info_list[0],":",info_list[1])
#     del info_list[0]
#     del info_list[0]
#     # print(info_list)
import openpyxl
wb = openpyxl.load_workbook('2023-02-21-23.15 Albums Month Data.xlsx')
ws = wb.active
for rowNum in range(2, ws.max_row):
    data = ws.cell(row=rowNum, column=3).value
    print(data)
    ws.cell(row=rowNum, column=3, value=data.replace("!)",'"'))
wb.save('2023-02-21-23.15 Albums Month Data.xlsx')
    # if state == 'OK':
    #     ws.cell(row=rowNum, column=3, value='In Inventory')
    # elif state == 'QNS':
    #     ws.cell(row=rowNum, column=3, value='(I am) Exhausted')