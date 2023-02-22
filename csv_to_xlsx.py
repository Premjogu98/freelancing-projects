import csv
from openpyxl import Workbook,load_workbook

# Variables
def change_something(xlsx_filepath):
    wb = load_workbook(xlsx_filepath)
    ws = wb.active
    for rowNum in range(2, ws.max_row):
        data = ws.cell(row=rowNum, column=3).value
        # print(data)
        ws.cell(row=rowNum, column=3, value=data.replace("!)",'"'))
    wb.save(xlsx_filepath)

def csv_to_xlsx(csv_filepath=None, xlsx_filepath=None,special=False):
    wb = Workbook()
    ws = wb.active
    if csv_filepath != None and xlsx_filepath != None:
        try:
            with open(csv_filepath, encoding="utf8") as f:
                reader = csv.reader(f)
                for row_index, row in enumerate(reader):
                    for column_index, cell in enumerate(row):
                        column_letter = column_index + 1
                        cell_value = cell.replace('"', '')
                        ws.cell(row = row_index + 1, column = column_letter).value = cell_value

                for row in ws.iter_rows(min_row=1, min_col=1, max_col=11, max_row=ws.max_row):
                    for cell in row:
                        if cell.value is None:
                            break
                        else:
                            try:
                                cell.value = float(cell.value)
                            except ValueError:
                                cell.value = cell.value
                        # print(type(cell.value))
                    wb.save(filename = xlsx_filepath)
            if special:
                change_something(xlsx_filepath)
            return True
        except Exception as e:
            print(e)
            return False
    else:
        print("FILE PATH MISSING PLEASE ENTER FILE PATH")
        return False

