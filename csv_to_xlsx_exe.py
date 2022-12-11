import csv
from openpyxl import Workbook

# Variables

def csv_to_xlsx(csv_filepath=None, xlsx_filepath=None):
    print("\n\n*** Welcome to File CSV to XLSX Converter EXE ***\n\n")
    wb = Workbook()
    ws = wb.active
    find = True
    if csv_filepath == None and xlsx_filepath == None:
        find = False
        csv_filepath = input("Please Enter CSV File Path/Name : ")
        if csv_filepath.strip() != "":
            xlsx_filepath = csv_filepath.replace(".csv",".xlsx")
            find = True
        else:
            print("NO FILE PATH OR NAME FOUND")
    if find:
        try:
            with open(csv_filepath, encoding="utf8") as f:
                print("File Conversion Processing.....\n")
                reader = csv.reader(f)
                for row_index, row in enumerate(reader):
                    for column_index, cell in enumerate(row):
                        column_letter = column_index + 1
                        cell_value = cell.replace('"', '')
                        ws.cell(row = row_index + 1, column = column_letter).value = cell_value

                for row in ws.iter_rows(min_row=1, min_col=1, max_col=11, max_row=ws.max_row):
                    for cell in row:
                        if cell.value is None:
                            break
                        else:
                            try:
                                cell.value = float(cell.value)
                            except ValueError:
                                cell.value = cell.value
                        # print(type(cell.value))
                    wb.save(filename = xlsx_filepath)
                print("File Processing Successfully Done ....\n")
                print(f"Your CSV File == {csv_filepath} CONVERTED Into XLSX File {xlsx_filepath}")
                return True
        except Exception as e:
            print(e)
            return False
csv_to_xlsx()